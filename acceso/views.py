import json
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

def _enviar_mailersend(destinatario_email, destinatario_nombre, asunto, texto_plano, html):
    import requests as http_requests
    from django.conf import settings
    try:
        response = http_requests.post(
            "https://api.mailersend.com/v1/email",
            headers={
                "Authorization": f"Bearer {settings.MAILERSEND_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "from": {
                    "email": settings.MAILERSEND_FROM_EMAIL,
                    "name": settings.MAILERSEND_FROM_NAME,
                },
                "to": [{"email": destinatario_email, "name": destinatario_nombre}],
                "subject": asunto,
                "text": texto_plano,
                "html": html,
            },
            timeout=10,
        )
        return response.status_code in (200, 202)
    except Exception:
        return False


# TIENDA PÚBLICA
def index(request):
    from .models import Sede

    sedes_index = Sede.objects.filter(activa=True).order_by("idSede")
    return render(request, "acceso/index.html", {"sedes_index": sedes_index})


def cerrar_sesion(request):
    logout(request)
    return redirect("index")

def login_view(request):
    if request.user.is_authenticated:
        return redirect("index")
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)
        if user is not None:
            from .models import Usuarios
            try:
                usuario_extra = Usuarios.objects.get(correoUsuario=user.email)
                if usuario_extra.estadoUsuario == "I":
                    messages.error(
                        request,
                        "⛔ Tu cuenta se encuentra inactiva. "
                        "Contacta al administrador: jz330777@gmail.com"
                    )
                    return render(request, "acceso/registro/login.html")
            except Usuarios.DoesNotExist:
                pass

            login(request, user)
            if user.is_staff:
                return redirect("panel_dashboard")
            if hasattr(user, "empleado"):
                return redirect("panel_empleado")
            return redirect("index")
        else:
            messages.error(request, "Usuario o contraseña incorrectos")
    return render(request, "acceso/registro/login.html")

def registro(request):
    if request.user.is_authenticated:
        return redirect("index")

    form_data = {}
    if request.method == "POST":
        from .models import Rol, Usuarios

        form_data = request.POST

        nombre = request.POST.get("nombreUsuario", "").strip()
        apellidos = request.POST.get("apellidosUsuario", "").strip()
        tipo_doc = request.POST.get("tipoDocUsuario", "CC")
        num_doc = request.POST.get("numDocUsuario", "").strip()
        correo = request.POST.get("correoUsuario", "").strip()
        telefono = request.POST.get("telefonoUsuario", "").strip()
        direccion = request.POST.get("direccionUsuario", "").strip()
        clave = request.POST.get("claveUsuario", "")
        clave2 = request.POST.get("claveUsuario2", "")

        if not nombre or not num_doc or not correo or not clave:
            messages.error(request, "Los campos marcados con * son obligatorios.")
            return render(
                request,
                "acceso/registro/registro.html",
                {"form_data": form_data},
            )

        if clave != clave2:
            messages.error(request, "Las contraseñas no coinciden.")
            return render(
                request,
                "acceso/registro/registro.html",
                {"form_data": form_data},
            )

        if len(clave) < 6:
            messages.error(request, "La contraseña debe tener al menos 6 caracteres.")
            return render(
                request,
                "acceso/registro/registro.html",
                {"form_data": form_data},
            )

        if Usuarios.objects.filter(numDocUsuario=num_doc).exists():
            messages.error(request, "Ya existe una cuenta con ese número de documento.")
            return render(
                request,
                "acceso/registro/registro.html",
                {"form_data": form_data},
            )

        if correo and Usuarios.objects.filter(correoUsuario=correo).exists():
            messages.error(request, "Ya existe una cuenta con ese correo.")
            return render(
                request,
                "acceso/registro/registro.html",
                {"form_data": form_data},
            )

        try:
            rol_cliente = Rol.objects.get(idRol=2)
        except Rol.DoesNotExist:
            rol_cliente = Rol.objects.first()

        from django.contrib.auth.hashers import make_password

        usuario = Usuarios(
            tipoDocUsuario=tipo_doc,
            numDocUsuario=num_doc,
            nombreUsuario=nombre,
            apellidosUsuario=apellidos if apellidos else None,
            correoUsuario=correo if correo else None,
            telefonoUsuario=int(telefono) if telefono else None,
            direccionUsuario=direccion if direccion else None,
            claveUsuario=make_password(clave),
            estadoUsuario="A",
            idRol=rol_cliente,
        )
        usuario.save()

        from django.contrib.auth.models import User

        # ✅ Toma el username del formulario
        username = request.POST.get("usernameUsuario", "").strip()

        if not username:
            messages.error(request, "El nombre de usuario es obligatorio.")
            return render(request, "acceso/registro/registro.html", {"form_data": form_data})

        if User.objects.filter(username=username).exists():
            messages.error(request, "Ese nombre de usuario ya está en uso.")
            return render(request, "acceso/registro/registro.html", {"form_data": form_data})

        django_user = User.objects.create_user(
            username=username,
            email=correo,
            password=clave,
            first_name=nombre,
            last_name=apellidos,
        )

        try:
            html_bienvenida = render_to_string("acceso/emails/bienvenida.html", {"nombre": nombre, "username": django_user.username})
            _enviar_mailersend(correo, nombre, "🏍️ ¡Bienvenido a Motopartes!", f"Hola {nombre}, tu cuenta ha sido creada.", html_bienvenida)
        except Exception:
            pass

        messages.success(request, f"¡Cuenta creada exitosamente! Bienvenido, {nombre}.")
        login(request, django_user)
        return redirect("index")

    return render(request, "acceso/registro/registro.html", {"form_data": form_data})


def productos(request):
    from .models import CategoriaProducto, Producto, Sede

    productos = Producto.objects.all()
    categorias = CategoriaProducto.objects.all()
    sedes_disponibles = Sede.objects.filter(activa=True)
    return render(
        request,
        "acceso/productos.html",
        {
            "productos": productos,
            "categorias": categorias,
            "sedes_disponibles": sedes_disponibles,
        },
    )


def detalle_producto(request, pk):
    from .models import CategoriaProducto, Producto, Sede

    producto = get_object_or_404(Producto, pk=pk)
    relacionados = Producto.objects.filter(
        idCategoria_Producto=producto.idCategoria_Producto
    ).exclude(pk=pk)[:4]
    categorias = CategoriaProducto.objects.all()
    sedes_disponibles = Sede.objects.filter(activa=True)
    return render(
        request,
        "acceso/detalle_producto.html",
        {
            "producto": producto,
            "relacionados": relacionados,
            "categorias": categorias,
            "sedes_disponibles": sedes_disponibles,
        },
    )


@login_required
def editar_producto(request, pk):
    from .models import CategoriaProducto, Producto

    producto = get_object_or_404(Producto, pk=pk)
    if not request.user.is_staff:
        return redirect("detalle_producto", pk=pk)
    if request.method == "POST":
        producto.nombreProducto = request.POST.get("nombreProducto")
        producto.precioProducto = request.POST.get("precioProducto")
        producto.stock = request.POST.get("stock")
        producto.estadoProducto = request.POST.get("estadoProducto")
        producto.descripcion = request.POST.get("descripcion")
        cat_id = request.POST.get("idCategoria_Producto")
        producto.idCategoria_Producto = CategoriaProducto.objects.get(pk=cat_id)
        if "imagen" in request.FILES:
            producto.imagen = request.FILES["imagen"]
        producto.save()
        return redirect("detalle_producto", pk=pk)
    return redirect("detalle_producto", pk=pk)


@login_required
def eliminar_producto(request, pk):
    from .models import Producto

    puede = request.user.is_staff or hasattr(request.user, "empleado")
    if puede and request.method == "POST":
        get_object_or_404(Producto, pk=pk).delete()
        messages.success(request, "Producto eliminado correctamente.")
    return redirect("productos")


# PANEL ADMIN
def es_admin(user):
    return user.is_staff or user.is_superuser


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_dashboard(request):
    from django.contrib.auth.models import User
    from django.utils import timezone

    from .models import CategoriaProducto, Pedido, Producto

    total_productos = Producto.objects.count()
    productos_agotados = Producto.objects.filter(stock=0).count()
    productos_stock_bajo = Producto.objects.filter(stock__gt=0, stock__lte=5).count()
    total_pedidos = Pedido.objects.count()
    pedidos_pendientes = Pedido.objects.filter(estadoPedido="Procesado").count()
    pedidos_hoy = Pedido.objects.filter(fechaPedido__date=timezone.now().date()).count()
    total_clientes = User.objects.filter(is_staff=False).count()
    total_categorias = CategoriaProducto.objects.count()

    ultimos_pedidos = Pedido.objects.select_related("idUsuario").order_by(
        "-fechaPedido"
    )[:6]
    productos_alerta = Producto.objects.filter(stock__lte=5).order_by("stock")[:8]

    return render(
        request,
        "acceso/panel/dashboard.html",
        {
            "total_productos": total_productos,
            "productos_agotados": productos_agotados,
            "productos_stock_bajo": productos_stock_bajo,
            "total_pedidos": total_pedidos,
            "pedidos_pendientes": pedidos_pendientes,
            "pedidos_hoy": pedidos_hoy,
            "total_clientes": total_clientes,
            "total_categorias": total_categorias,
            "ultimos_pedidos": ultimos_pedidos,
            "productos_alerta": productos_alerta,
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_productos(request):
    from .models import CategoriaProducto, Producto

    q = request.GET.get("q", "")
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")

    productos = Producto.objects.select_related("idCategoria_Producto").order_by(
        "nombreProducto"
    )

    if q:
        productos = productos.filter(nombreProducto__icontains=q)
    if categoria_id:
        productos = productos.filter(
            idCategoria_Producto__idCategoria_Producto=categoria_id
        )
    if estado == "Agotado":
        productos = productos.filter(stock=0)
    elif estado == "stock_bajo":
        productos = productos.filter(stock__gt=0, stock__lte=5)
    elif estado:
        productos = productos.filter(estadoProducto=estado)

    categorias = CategoriaProducto.objects.all()
    return render(
        request,
        "acceso/panel/panel_productos.html",
        {
            "productos": productos,
            "categorias": categorias,
            "q": q,
            "categoria_id": categoria_id,
            "estado": estado,
            "total": productos.count(),
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_producto_crear(request):
    from .models import CategoriaProducto, Producto

    categorias = CategoriaProducto.objects.all()
    if request.method == "POST":
        producto = Producto(
            nombreProducto=request.POST.get("nombreProducto"),
            precioProducto=request.POST.get("precioProducto"),
            stock=request.POST.get("stock", 0),
            estadoProducto=request.POST.get("estadoProducto", "Disponible"),
            descripcion=request.POST.get("descripcion", ""),
        )
        cat_id = request.POST.get("idCategoria_Producto")
        if cat_id:
            producto.idCategoria_Producto = CategoriaProducto.objects.get(pk=cat_id)
        if "imagen" in request.FILES:
            producto.imagen = request.FILES["imagen"]
        producto.save()
        messages.success(
            request,
            f'Producto "{producto.nombreProducto}" creado correctamente.',
        )
        return redirect("panel_productos")
    return render(
        request,
        "acceso/panel/panel_producto_form.html",
        {"titulo": "Nuevo Producto", "categorias": categorias},
        
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_producto_editar(request, pk):
    from .models import CategoriaProducto, Producto

    producto = get_object_or_404(Producto, pk=pk)
    categorias = CategoriaProducto.objects.all()
    if request.method == "POST":
        producto.nombreProducto = request.POST.get("nombreProducto")
        producto.precioProducto = request.POST.get("precioProducto")
        producto.stock = request.POST.get("stock", 0)
        producto.estadoProducto = request.POST.get("estadoProducto", "Disponible")
        producto.descripcion = request.POST.get("descripcion", "")
        cat_id = request.POST.get("idCategoria_Producto")
        if cat_id:
            producto.idCategoria_Producto = CategoriaProducto.objects.get(pk=cat_id)
        if "imagen" in request.FILES:
            producto.imagen = request.FILES["imagen"]
        producto.save()
        messages.success(request, f'Producto "{producto.nombreProducto}" actualizado.')
        return redirect("panel_productos")
    return render(
        request,
        "acceso/panel/panel_producto_form.html",
        {
            "titulo": "Editar Producto",
            "producto": producto,
            "categorias": categorias,
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_producto_eliminar(request, pk):
    from .models import Producto

    producto = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        nombre = producto.nombreProducto
        producto.delete()
        messages.success(request, f'Producto "{nombre}" eliminado.')
        return redirect("panel_productos")
    return render(
        request,
        "acceso/panel/panel_confirmar_eliminar.html",
        {"objeto": producto},
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_pedidos(request):
    from .models import Pedido

    q = request.GET.get("q", "")
    estado_filtro = request.GET.get("estado", "")

    pedidos = Pedido.objects.select_related("idUsuario").order_by("-fechaPedido")

    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q) | Q(ciudad__icontains=q)
        )
    if estado_filtro:
        pedidos = pedidos.filter(estadoPedido=estado_filtro)

    return render(
        request,
        "acceso/panel/panel_pedidos.html",
        {
            "pedidos": pedidos,
            "q": q,
            "estado_filtro": estado_filtro,
            "total": pedidos.count(),
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_pedido_detalle(request, pk):
    from .models import Pedido

    pedido = get_object_or_404(Pedido.objects.select_related("idUsuario"), pk=pk)
    if request.method == "POST":
        pedido.estadoPedido = request.POST.get("estadoPedido", pedido.estadoPedido)
        pedido.save()
        messages.success(request, f"Estado del pedido #{pk} actualizado.")
        return redirect("panel_pedido_detalle", pk=pk)
    return render(request, "acceso/panel/panel_pedido_detalle.html", {"pedido": pedido})

@login_required
@user_passes_test(es_admin, login_url="/")
def panel_clientes(request):
    from django.contrib.auth.models import User
    from .models import Usuarios

    q = request.GET.get("q", "")
    clientes = Usuarios.objects.annotate(num_pedidos=Count("pedido")).order_by(
        "nombreUsuario"
    )
    if q:
        clientes = clientes.filter(
            Q(nombreUsuario__icontains=q)
            | Q(correoUsuario__icontains=q)
            | Q(numDocUsuario__icontains=q)
        )

     for cliente in clientes:
        django_user = User.objects.filter(email=cliente.correoUsuario).first()
        cliente.username_django = django_user.username if django_user else "—"

    return render(
        request,
        "acceso/panel/panel_clientes.html",
        {"clientes": clientes, "q": q},
    )

def login_empleado(request):
    if request.user.is_authenticated and hasattr(request.user, "empleado"):
        return redirect("panel_empleado")
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user and hasattr(user, "empleado"):
            login(request, user)
            return redirect("panel_empleado")
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")
    return render(request, "acceso/registro/login.html")


@login_required
def panel_empleado(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")

    from .models import Producto

    productos = Producto.objects.all().order_by("nombreProducto")

    return render(
        request,
        "acceso/empleado/empleado_dashboard.html",
        {
            "productos": productos,
            "total": productos.count(),
        },
    )


@login_required
def empleado_producto_crear(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import CategoriaProducto, Producto

    categorias = CategoriaProducto.objects.all()
    if request.method == "POST":
        nombre = request.POST.get("nombreProducto", "").strip()
        precio = request.POST.get("precioProducto", "0")
        stock = request.POST.get("stock", "0")
        estado = request.POST.get("estadoProducto", "Disponible")
        desc = request.POST.get("descripcion", "").strip()
        cat_id = request.POST.get("idCategoria_Producto")
        imagen = request.FILES.get("imagen")
        if not nombre:
            messages.error(request, "El nombre es obligatorio.")
        else:
            cat = (
                CategoriaProducto.objects.filter(pk=cat_id).first()
                if cat_id and cat_id.strip()
                else None
            )
            p = Producto(
                nombreProducto=nombre,
                precioProducto=precio or 0,
                stock=stock or 0,
                estadoProducto=estado,
                descripcion=desc,
                idCategoria_Producto=cat,
            )
            if imagen:
                p.imagen = imagen
            p.save()
            messages.success(request, f'Producto "{nombre}" creado.')
            return redirect("panel_empleado")
    return render(
        request,
        "acceso/empleado/empleado_producto_form.html",
        {"titulo": "Nuevo Producto", "categorias": categorias},
    )


@login_required
def empleado_producto_editar(request, pk):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import CategoriaProducto, Producto

    producto = get_object_or_404(Producto, pk=pk)
    categorias = CategoriaProducto.objects.all()
    if request.method == "POST":
        producto.nombreProducto = request.POST.get("nombreProducto", "").strip()
        producto.precioProducto = request.POST.get("precioProducto", "0")
        producto.stock = request.POST.get("stock", "0")
        producto.estadoProducto = request.POST.get("estadoProducto", "Disponible")
        producto.descripcion = request.POST.get("descripcion", "").strip()
        cat_id = request.POST.get("idCategoria_Producto")
        producto.idCategoria_Producto = (
            CategoriaProducto.objects.filter(pk=cat_id).first()
            if cat_id and cat_id.strip()
            else None
        )
        if request.FILES.get("imagen"):
            producto.imagen = request.FILES.get("imagen")
        producto.save()
        messages.success(request, "Producto actualizado.")
        return redirect("panel_empleado")
    return render(
        request,
        "acceso/empleado/empleado_producto_form.html",
        {
            "titulo": "Editar Producto",
            "producto": producto,
            "categorias": categorias,
        },
    )


@login_required
def empleado_producto_eliminar(request, pk):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import Producto

    producto = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        nombre = producto.nombreProducto
        producto.delete()
        messages.success(request, f'Producto "{nombre}" eliminado.')
        return redirect("panel_empleado")
    return render(
        request,
        "acceso/empleado/empleado_confirmar_eliminar.html",
        {"producto": producto},
    )


@login_required
def agregar_producto_empleado(request):
    return redirect("empleado_producto_crear")


# PANEL EMPLEADOS (admin gestiona empleados)


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_empleados(request):
    from .models import Empleado

    q = request.GET.get("q", "")
    empleados = Empleado.objects.select_related("usuario").order_by(
        "usuario__first_name"
    )
    if q:
        empleados = empleados.filter(
            Q(usuario__first_name__icontains=q)
            | Q(usuario__last_name__icontains=q)
            | Q(cedula__icontains=q)
            | Q(cargo__icontains=q)
        )
    return render(
        request,
        "acceso/panel/panel_empleados.html",
        {"empleados": empleados, "q": q, "total": empleados.count()},
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_empleado_crear(request):
    from django.contrib.auth.models import User

    from .models import Empleado

    cargos = [("empleado", "Empleado"), ("admin", "Administrador")]
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        cedula = request.POST.get("cedula", "").strip()
        cargo = request.POST.get("cargo", "empleado")
        telefono = request.POST.get("telefono", "").strip()
        direccion = request.POST.get("direccion", "").strip()
        fnac = request.POST.get("fecha_nacimiento", "").strip()
        salario = request.POST.get("salario", "0").strip()

        errores = []
        if not first_name:
            errores.append("El nombre es obligatorio.")
        if not cedula:
            errores.append("La cédula es obligatoria.")
        if not username:
            errores.append("El usuario es obligatorio.")
        if not password:
            errores.append("La contraseña es obligatoria.")
        if not fnac:
            errores.append("La fecha de nacimiento es obligatoria.")
        if User.objects.filter(username=username).exists():
            errores.append("Ese nombre de usuario ya existe.")
        if Empleado.objects.filter(cedula=cedula).exists():
            errores.append("Ya existe un empleado con esa cédula.")

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=(cargo == "admin"),
            )
            Empleado.objects.create(
                usuario=user,
                cedula=cedula,
                cargo=cargo,
                telefono=telefono,
                direccion=direccion,
                fecha_nacimiento=fnac,
                salario=salario or 0,
            )
            messages.success(request, f"Empleado {first_name} {last_name} creado.")
            return redirect("panel_empleados")
    return render(
        request,
        "acceso/panel/panel_empleado_form.html",
        {
            "titulo": "Nuevo Empleado",
            "cargos": cargos,
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_empleado_editar(request, pk):
    from .models import Empleado

    cargos = [("empleado", "Empleado"), ("admin", "Administrador")]
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == "POST":
        fnac = request.POST.get("fecha_nacimiento", "").strip()
        if not fnac:
            messages.error(request, "La fecha de nacimiento es obligatoria.")
            return render(
                request,
                "acceso/panel/panel_empleado_form.html",
                {
                    "titulo": "Editar Empleado",
                    "empleado": empleado,
                    "cargos": cargos,
                },
            )
        cargo = request.POST.get("cargo", "empleado")
        empleado.usuario.first_name = request.POST.get("first_name", "").strip()
        empleado.usuario.last_name = request.POST.get("last_name", "").strip()
        empleado.usuario.email = request.POST.get("email", "").strip()
        empleado.usuario.is_staff = cargo == "admin"
        empleado.usuario.save()
        empleado.cedula = request.POST.get("cedula", "").strip()
        empleado.cargo = cargo
        empleado.telefono = request.POST.get("telefono", "").strip()
        empleado.direccion = request.POST.get("direccion", "").strip()
        empleado.fecha_nacimiento = fnac
        empleado.salario = request.POST.get("salario", "0").strip() or 0
        empleado.save()
        messages.success(request, "Empleado actualizado.")
        return redirect("panel_empleados")
    return render(
        request,
        "acceso/panel/panel_empleado_form.html",
        {"titulo": "Editar Empleado", "empleado": empleado, "cargos": cargos},
    )



@login_required
@user_passes_test(es_admin, login_url="/")
def panel_empleado_eliminar(request, pk):
    from .models import Empleado

    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == "POST":
        nombre = empleado.usuario.get_full_name()
        empleado.usuario.delete()
        messages.success(request, f"Empleado {nombre} eliminado.")
        return redirect("panel_empleados")
    return render(
        request,
        "acceso/panel/panel_confirmar_eliminar.html",
        {"objeto": empleado},
    )


# PANEL CATEGORÍAS
@login_required
@user_passes_test(es_admin, login_url="/")
def panel_categorias(request):
    from .models import CategoriaProducto

    q = request.GET.get("q", "")
    categorias = CategoriaProducto.objects.annotate(
        num_productos=Count("producto")
    ).order_by("nombreCategoria")
    if q:
        categorias = categorias.filter(
            Q(nombreCategoria__icontains=q) | Q(descripcion__icontains=q)
        )
    return render(
        request,
        "acceso/panel/panel_categorias.html",
        {"categorias": categorias, "q": q},
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_categoria_crear(request):
    from .models import CategoriaProducto

    if request.method == "POST":
        nombre = request.POST.get("nombreCategoria", "").strip()
        desc = request.POST.get("descripcion", "").strip()
        if not nombre:
            messages.error(request, "El nombre es obligatorio.")
        else:
            CategoriaProducto.objects.create(nombreCategoria=nombre, descripcion=desc)
            messages.success(request, f'Categoría "{nombre}" creada.')
            return redirect("panel_categorias")
    return render(
        request,
        "acceso/panel/panel_categoria_form.html",
        {"titulo": "Nueva Categoría"},
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_categoria_editar(request, pk):
    from .models import CategoriaProducto

    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    if request.method == "POST":
        categoria.nombreCategoria = request.POST.get("nombreCategoria", "").strip()
        categoria.descripcion = request.POST.get("descripcion", "").strip()
        categoria.save()
        messages.success(request, "Categoría actualizada.")
        return redirect("panel_categorias")
    return render(
        request,
        "acceso/panel/panel_categoria_form.html",
        {"titulo": "Editar Categoría", "categoria": categoria},
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_categoria_eliminar(request, pk):
    from .models import CategoriaProducto

    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    if request.method == "POST":
        nombre = categoria.nombreCategoria
        categoria.delete()
        messages.success(request, f'Categoría "{nombre}" eliminada.')
        return redirect("panel_categorias")
    return render(
        request,
        "acceso/panel/panel_confirmar_eliminar.html",
        {"objeto": categoria},
    )


# PANEL REPORTES
@login_required
@user_passes_test(es_admin, login_url="/")
def panel_reportes(request):
    from django.db.models import Sum

    from .models import CategoriaProducto, Empleado, Pedido, Producto, Usuarios

    total_ventas = Pedido.objects.aggregate(total=Sum("totalPedido"))["total"] or 0
    total_pedidos = Pedido.objects.count()
    pedidos_pend = Pedido.objects.filter(estadoPedido="Procesado").count()
    pedidos_entregados = Pedido.objects.filter(estadoPedido="Entregado").count()
    total_productos = Producto.objects.count()
    sin_stock = Producto.objects.filter(stock=0).count()
    stock_bajo = Producto.objects.filter(stock__gt=0, stock__lte=5).count()
    total_clientes = Usuarios.objects.count()
    total_empleados = Empleado.objects.count()
    total_categorias = CategoriaProducto.objects.count()
    productos_top = Producto.objects.order_by("-stock")[:5]
    ultimos_pedidos = Pedido.objects.select_related("idUsuario").order_by(
        "-fechaPedido"
    )[:10]

    return render(
        request,
        "acceso/panel/panel_reportes.html",
        {
            "total_ventas": total_ventas,
            "total_pedidos": total_pedidos,
            "pedidos_pend": pedidos_pend,
            "pedidos_entregados": pedidos_entregados,
            "total_productos": total_productos,
            "sin_stock": sin_stock,
            "stock_bajo": stock_bajo,
            "total_clientes": total_clientes,
            "total_empleados": total_empleados,
            "total_categorias": total_categorias,
            "productos_top": productos_top,
            "ultimos_pedidos": ultimos_pedidos,
        },
    )


# REPORTES PDF
def _generar_pdf(html_string):
    import io

    from django.http import HttpResponse
    from xhtml2pdf import pisa

    buffer = io.BytesIO()
    resultado = pisa.CreatePDF(html_string, dest=buffer)
    if resultado.err:
        return HttpResponse("Error al generar el PDF", status=500)
    buffer.seek(0)
    return buffer


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_empleados_pdf(request):
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import Empleado

    empleados = Empleado.objects.select_related("usuario").order_by(
        "usuario__first_name"
    )
    html = render_to_string(
        "acceso/empleado/pdf/reporte_empleados_pdf.html",
        {
            "empleados": empleados,
            "fecha": timezone.now(),
            "total": empleados.count(),
        },
    )
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_empleados.pdf"'
    return response


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_productos_pdf(request):
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import Producto

    productos = Producto.objects.select_related("idCategoria_Producto").order_by(
        "nombreProducto"
    )
    html = render_to_string(
        "acceso/empleado/pdf/reporte_productos_pdf.html",
        {
            "productos": productos,
            "fecha": timezone.now(),
            "total": productos.count(),
        },
    )
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_productos.pdf"'
    return response


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_pedidos_pdf(request):
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import Pedido

    pedidos = Pedido.objects.select_related("idUsuario").order_by("-fechaPedido")
    html = render_to_string(
        "acceso/empleado/pdf/reporte_pedidos_pdf.html",
        {
            "pedidos": pedidos,
            "fecha": timezone.now(),
            "total": pedidos.count(),
        },
    )
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_pedidos.pdf"'
    return response


# CARRITO DE COMPRAS
def _get_usuario_moto(django_user):
    from .models import Usuarios

    try:
        return Usuarios.objects.get(correoUsuario=django_user.email)
    except Usuarios.DoesNotExist:
        return None


@login_required
def ver_carrito(request):
    from .models import CarritoItem, Sede

    items = CarritoItem.objects.filter(usuario=request.user).select_related(
        "idProducto"
    )
    total = sum(item.subtotal for item in items)
    sedes = Sede.objects.filter(activa=True)
    return render(
        request,
        "acceso/carrito.html",
        {"items": items, "total": total, "sedes": sedes},
    )


@login_required
def agregar_al_carrito(request, pk):
    from .models import CarritoItem, Producto

    if request.method == "POST":
        producto = get_object_or_404(Producto, pk=pk)
        if producto.stock <= 0:
            messages.warning(request, "Este producto está agotado.")
            return redirect("detalle_producto", pk=pk)
        cantidad = max(1, int(request.POST.get("cantidad", 1)))
        item, creado = CarritoItem.objects.get_or_create(
            usuario=request.user,
            idProducto=producto,
            defaults={"cantidad": cantidad},
        )
        if not creado:
            item.cantidad = min(item.cantidad + cantidad, producto.stock)
            item.save()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            from django.http import JsonResponse

            total_items = CarritoItem.objects.filter(usuario=request.user).count()
            return JsonResponse({"ok": True, "total_items": total_items})
        messages.success(request, f'"{producto.nombreProducto}" agregado al carrito.')
        return redirect("ver_carrito")
    return redirect("productos")


@login_required
def actualizar_cantidad_carrito(request, pk):
    from .models import CarritoItem

    if request.method == "POST":
        item = get_object_or_404(CarritoItem, pk=pk, usuario=request.user)
        cantidad = int(request.POST.get("cantidad", 1))
        if cantidad < 1:
            item.delete()
            messages.info(request, "Producto eliminado del carrito.")
        else:
            item.cantidad = min(cantidad, item.idProducto.stock)
            item.save()
    return redirect("ver_carrito")


@login_required
def eliminar_del_carrito(request, pk):
    from .models import CarritoItem

    item = get_object_or_404(CarritoItem, pk=pk, usuario=request.user)
    nombre = item.idProducto.nombreProducto
    item.delete()
    messages.info(request, f'"{nombre}" eliminado del carrito.')
    return redirect("ver_carrito")


@login_required
def checkout(request):
    from decimal import Decimal

    from .models import CarritoItem, DetallePedido, Factura, Pedido, Sede

    COSTOS_ENVIO = {
        "Bogotá": Decimal("5000"),
        "Medellín": Decimal("8000"),
        "Cali": Decimal("8000"),
        "Barranquilla": Decimal("10000"),
    }

    items = CarritoItem.objects.filter(usuario=request.user).select_related(
        "idProducto"
    )

    if not items.exists():
        messages.warning(request, "Tu carrito está vacío.")
        return redirect("ver_carrito")

    sedes = Sede.objects.filter(activa=True)

    if request.method == "POST":
        metodo = request.POST.get("metodoPago", "Efectivo")
        tipo_entrega = request.POST.get("tipo_entrega", "sede")

        if tipo_entrega == "sede":
            sede_id = request.POST.get("sede")
            if not sede_id:
                messages.error(request, "Debes seleccionar una sede de recogida.")
                total = sum(item.subtotal for item in items)
                return render(
                    request,
                    "acceso/carrito.html",
                    {"items": items, "total": total, "sedes": sedes},
                )
            sede = get_object_or_404(Sede, pk=sede_id, activa=True)
            direccion_dom = barrio_dom = ciudad_dom = telefono_dom = None
            costo_envio = Decimal("0")
            ciudad_pedido = sede.ciudad
        else:
            ciudad_dom = request.POST.get("ciudad_domicilio", "").strip()
            direccion_dom = request.POST.get("direccion_domicilio", "").strip()
            barrio_dom = request.POST.get("barrio_domicilio", "").strip()
            telefono_dom = request.POST.get("telefono_domicilio", "").strip()

            errores = []
            if not ciudad_dom:
                errores.append("Debes seleccionar una ciudad de entrega.")
            if not direccion_dom:
                errores.append("La dirección de entrega es obligatoria.")
            if not barrio_dom:
                errores.append("El barrio es obligatorio.")
            if not telefono_dom:
                errores.append("El teléfono de contacto es obligatorio.")

            if errores:
                for e in errores:
                    messages.error(request, e)
                total = sum(item.subtotal for item in items)
                return render(
                    request,
                    "acceso/carrito.html",
                    {"items": items, "total": total, "sedes": sedes},
                )

            sede = None
            costo_envio = COSTOS_ENVIO.get(ciudad_dom, Decimal("0"))
            ciudad_pedido = ciudad_dom

        usuario_moto = _get_usuario_moto(request.user)
        if not usuario_moto:
            messages.error(
                request, "No se encontró tu perfil. Contacta al administrador."
            )
            return redirect("ver_carrito")

        subtotal = sum(item.subtotal for item in items)
        iva = round(subtotal * Decimal("0.19"), 2)
        total = subtotal + iva + costo_envio

        pedido = Pedido.objects.create(
            idUsuario=usuario_moto,
            sede=sede,
            estadoPedido="Procesado",
            metodoPago=metodo,
            totalPedido=total,
            codigopostal="",
            ciudad=ciudad_pedido,
            tipo_entrega=tipo_entrega,
            direccion_domicilio=direccion_dom,
            barrio_domicilio=barrio_dom,
            ciudad_domicilio=ciudad_dom,
            telefono_domicilio=telefono_dom,
            costo_envio=costo_envio,
        )

        for item in items:
            DetallePedido.objects.create(
                idPedido=pedido,
                idProducto=item.idProducto,
                cantidad=item.cantidad,
                precioUnitario=item.idProducto.precioProducto,
            )
            prod = item.idProducto
            prod.stock = max(0, prod.stock - item.cantidad)
            if prod.stock == 0:
                prod.estadoProducto = "Agotado"
            prod.save(update_fields=["stock", "estadoProducto"])

        ultimo = Factura.objects.order_by("-idFactura").first()
        siguiente_num = (ultimo.idFactura + 1) if ultimo else 1
        numero_factura = f"FAC-{siguiente_num:06d}"

        factura = Factura.objects.create(
            idPedido=pedido,
            tipoDocUsuario=usuario_moto.tipoDocUsuario,
            numeroDoc=usuario_moto.numDocUsuario,
            valorUnitarioProducto=subtotal,
            subtotal=subtotal,
            iva=iva,
            totalPedido=total,
            fechaPedido=pedido.fechaPedido,
            metodoDePago=metodo,
            numeroFactura=numero_factura,
        )

        items.delete()

        try:
            detalles_pedido = pedido.detallepedido_set.select_related("idProducto").all()
            html_correo = render_to_string("acceso/emails/confirmacion_pedido.html", {"nombre_cliente": usuario_moto.nombreUsuario, "pedido": pedido, "factura": factura, "detalles": detalles_pedido})
            _enviar_mailersend(request.user.email, usuario_moto.nombreUsuario, f"Pedido #{pedido.idPedido} confirmado — Motopartes", f"Tu pedido #{pedido.idPedido} ha sido confirmado.", html_correo)
        except Exception:
            pass
        
        messages.success(
            request,
            f"¡Pedido #{pedido.idPedido} confirmado!"
            f"Factura {numero_factura} generada.",
        )
        return redirect("ver_factura", pk=factura.idFactura)

    total = sum(item.subtotal for item in items)
    return render(
        request,
        "acceso/carrito.html",
        {"items": items, "total": total, "sedes": sedes},
    )


@login_required
def mis_facturas(request):
    from .models import Factura

    usuario_moto = _get_usuario_moto(request.user)
    if not usuario_moto:
        messages.error(request, "No se encontró tu perfil.")
        return redirect("index")
    facturas = (
        Factura.objects.filter(idPedido__idUsuario=usuario_moto)
        .select_related("idPedido", "idPedido__sede")
        .order_by("-fechaFactura")
    )
    return render(request, "acceso/mis_compras.html", {"facturas": facturas})


@login_required
def detalle_factura(request, pk):
    from .models import DetallePedido, Factura

    usuario_moto = _get_usuario_moto(request.user)
    if not usuario_moto:
        return redirect("index")
    factura = get_object_or_404(Factura, pk=pk, idPedido__idUsuario=usuario_moto)
    detalles = DetallePedido.objects.filter(idPedido=factura.idPedido).select_related(
        "idProducto"
    )
    return render(
        request,
        "acceso/factura_detalle.html",
        {"factura": factura, "detalles": detalles},
    )


# PANEL EMPLEADO — PEDIDOS
@login_required
def empleado_pedidos(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import Pedido, Sede

    estado = request.GET.get("estado", "")
    sede_id = request.GET.get("sede", "")
    pedidos = Pedido.objects.select_related("idUsuario", "sede").order_by(
        "-fechaPedido"
    )
    if estado:
        pedidos = pedidos.filter(estadoPedido=estado)
    if sede_id:
        pedidos = pedidos.filter(sede_id=sede_id)
    sedes = Sede.objects.filter(activa=True)
    estados = [
        "Procesado",
        "Confirmado",
        "En Proceso",
        "Listo",
        "Entregado",
        "Cancelado",
    ]
    return render(
        request,
        "acceso/empleado/empleado_pedidos.html",
        {
            "pedidos": pedidos,
            "sedes": sedes,
            "estados": estados,
            "estado_actual": estado,
            "sede_actual": sede_id,
        },
    )


@login_required
def empleado_pedido_detalle(request, pk):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import DetallePedido, Pedido

    pedido = get_object_or_404(Pedido, pk=pk)
    detalles = DetallePedido.objects.filter(idPedido=pedido).select_related(
        "idProducto"
    )
    factura = getattr(pedido, "factura", None)
    if request.method == "POST":
        pedido.estadoPedido = request.POST.get("estadoPedido", pedido.estadoPedido)
        pedido.save(update_fields=["estadoPedido"])
        messages.success(
            request,
            f'Estado del pedido #{pk} actualizado a "{pedido.estadoPedido}".',
        )
        return redirect("empleado_pedido_detalle", pk=pk)
    estados = [
        "Procesado",
        "Confirmado",
        "En Proceso",
        "Listo",
        "Entregado",
        "Cancelado",
    ]
    return render(
        request,
        "acceso/empleado/empleado_pedido_detalle.html",
        {
            "pedido": pedido,
            "detalles": detalles,
            "factura": factura,
            "estados": estados,
        },
    )


# PANEL EMPLEADO — CATEGORÍAS
@login_required
def empleado_categorias(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from django.db.models import Count

    from .models import CategoriaProducto

    categorias = CategoriaProducto.objects.annotate(
        num_productos=Count("producto")
    ).order_by("nombreCategoria")
    return render(
        request,
        "acceso/empleado/empleado_categorias.html",
        {
            "categorias": categorias,
            "total": categorias.count(),
        },
    )


@login_required
def empleado_categoria_crear(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import CategoriaProducto

    if request.method == "POST":
        nombre = request.POST.get("nombreCategoria", "").strip()
        desc = request.POST.get("descripcion", "").strip()
        if not nombre:
            messages.error(request, "El nombre es obligatorio.")
        elif CategoriaProducto.objects.filter(nombreCategoria=nombre).exists():
            messages.error(request, f'Ya existe una categoría llamada "{nombre}".')
        else:
            CategoriaProducto.objects.create(nombreCategoria=nombre, descripcion=desc)
            messages.success(request, f'Categoría "{nombre}" creada correctamente.')
            return redirect("empleado_categorias")
    return render(
        request,
        "acceso/empleado/empleado_categoria_form.html",
        {"titulo": "Nueva Categoría"},
    )


@login_required
def empleado_categoria_editar(request, pk):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import CategoriaProducto

    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    if request.method == "POST":
        nombre = request.POST.get("nombreCategoria", "").strip()
        desc = request.POST.get("descripcion", "").strip()
        if not nombre:
            messages.error(request, "El nombre es obligatorio.")
        else:
            categoria.nombreCategoria = nombre
            categoria.descripcion = desc
            categoria.save()
            messages.success(request, f'Categoría "{nombre}" actualizada.')
            return redirect("empleado_categorias")
    return render(
        request,
        "acceso/empleado/empleado_categoria_form.html",
        {"titulo": "Editar Categoría", "categoria": categoria},
    )


@login_required
def empleado_categoria_eliminar(request, pk):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import CategoriaProducto

    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    if request.method == "POST":
        nombre = categoria.nombreCategoria
        categoria.delete()
        messages.success(request, f'Categoría "{nombre}" eliminada.')
        return redirect("empleado_categorias")
    return render(
        request,
        "acceso/empleado/empleado_confirmar_eliminar.html",
        {"categoria": categoria},
    )


# PANEL ADMIN — SEDES
@login_required
@user_passes_test(es_admin, login_url="/")
def panel_sedes(request):
    from django.db.models import Count

    from .models import Sede

    q = request.GET.get("q", "")
    sedes = Sede.objects.annotate(num_pedidos=Count("pedido")).order_by("nombre")
    if q:
        sedes = sedes.filter(
            Q(nombre__icontains=q) | Q(ciudad__icontains=q) | Q(direccion__icontains=q)
        )
    return render(
        request,
        "acceso/panel/panel_sedes.html",
        {
            "sedes": sedes,
            "total": sedes.count(),
            "q": q,
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_sede_crear(request):
    from .models import Sede

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        direccion = request.POST.get("direccion", "").strip()
        ciudad = request.POST.get("ciudad", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        activa = request.POST.get("activa") == "on"
        if not nombre or not direccion or not ciudad:
            messages.error(request, "Nombre, dirección y ciudad son obligatorios.")
        else:
            Sede.objects.create(
                nombre=nombre,
                direccion=direccion,
                ciudad=ciudad,
                telefono=telefono,
                activa=activa,
            )
            messages.success(request, f'Sede "{nombre}" creada correctamente.')
            return redirect("panel_sedes")
    return render(
        request, "acceso/panel/panel_sede_form.html", {"titulo": "Nueva Sede"}
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_sede_editar(request, pk):
    from .models import Sede

    sede = get_object_or_404(Sede, pk=pk)
    if request.method == "POST":
        sede.nombre = request.POST.get("nombre", "").strip()
        sede.direccion = request.POST.get("direccion", "").strip()
        sede.ciudad = request.POST.get("ciudad", "").strip()
        sede.telefono = request.POST.get("telefono", "").strip()
        sede.activa = request.POST.get("activa") == "on"
        if not sede.nombre or not sede.direccion or not sede.ciudad:
            messages.error(request, "Nombre, dirección y ciudad son obligatorios.")
        else:
            sede.save()
            messages.success(request, f'Sede "{sede.nombre}" actualizada.')
            return redirect("panel_sedes")
    return render(
        request,
        "acceso/panel/panel_sede_form.html",
        {"titulo": "Editar Sede", "sede": sede},
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_sede_eliminar(request, pk):
    from .models import Sede

    sede = get_object_or_404(Sede, pk=pk)
    if request.method == "POST":
        nombre = sede.nombre
        sede.delete()
        messages.success(request, f'Sede "{nombre}" eliminada.')
        return redirect("panel_sedes")
    return render(
        request, "acceso/panel/panel_confirmar_eliminar.html", {"objeto": sede}
    )


@login_required
@require_POST
def realizar_compra(request):
    from django.http import JsonResponse

    from .models import (
        DetallePedido,
        Factura,
        Pedido,
        Producto,
        Sede,
        Usuarios,
    )

    try:
        data = json.loads(request.body)
        producto_id = data.get("idProducto")
        sede_id = data.get("idSede")
        cantidad = int(data.get("cantidad", 1))
        metodo_pago = data.get("metodoPago", "Efectivo")

        producto = get_object_or_404(Producto, pk=producto_id)
        sede = get_object_or_404(Sede, pk=sede_id)

        if producto.stock < cantidad:
            return JsonResponse(
                {
                    "ok": False,
                    "error": f"Solo hay {producto.stock} unidades disponibles.",
                }
            )

        precio_unit = producto.precioProducto
        subtotal = precio_unit * cantidad
        iva = round(subtotal * Decimal("0.19"), 2)
        total = subtotal + iva

        producto.stock -= cantidad
        if producto.stock == 0:
            producto.estadoProducto = "Agotado"
        producto.save()

        try:
            usuario_custom = Usuarios.objects.get(correoUsuario=request.user.email)

        except Usuarios.DoesNotExist:
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "No se encontró tu perfil. " "Verifica tu correo de registro."
                    ),
                }
            )

        pedido = Pedido.objects.create(
            idUsuario=usuario_custom,
            sede=sede,
            metodoPago=metodo_pago,
            totalPedido=total,
            codigopostal="",
            ciudad=sede.ciudad,
        )
        DetallePedido.objects.create(
            idPedido=pedido,
            idProducto=producto,
            cantidad=cantidad,
            precioUnitario=precio_unit,
        )
        factura = Factura.objects.create(
            idPedido=pedido,
            tipoDocUsuario=usuario_custom.tipoDocUsuario,
            numeroDoc=usuario_custom.numDocUsuario,
            subtotal=subtotal,
            iva=iva,
            totalPedido=total,
            fechaPedido=pedido.fechaPedido,
            metodoDePago=metodo_pago,
        )

        return JsonResponse(
            {
                "ok": True,
                "idPedido": pedido.idPedido,
                "numeroFactura": factura.numeroFactura,
                "total": str(total),
                "iva": str(iva),
                "subtotal": str(subtotal),
                "fechaFactura": factura.fechaFactura.strftime("%d/%m/%Y %H:%M"),
                "sede": str(sede),
                "producto": producto.nombreProducto,
                "cantidad": cantidad,
                "metodoPago": metodo_pago,
                "cliente": request.user.get_full_name() or request.user.username,
            }
        )
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})


@login_required
def mis_compras(request):
    from .models import Pedido

    usuario_moto = _get_usuario_moto(request.user)
    pedidos = (
        Pedido.objects.filter(idUsuario=usuario_moto)
        .select_related("sede")
        .order_by("-fechaPedido")
    )
    return render(request, "acceso/mis_compras.html", {"pedidos": pedidos})


@login_required
def ver_factura(request, pk):
    from .models import Factura

    factura = get_object_or_404(
        Factura.objects.select_related("idPedido__sede", "idPedido__idUsuario"),
        pk=pk,
    )
    pedido = factura.idPedido
    usuario_moto = _get_usuario_moto(request.user)
    if (
        factura.idPedido.idUsuario != usuario_moto
        and not request.user.is_staff
        and not hasattr(request.user, "empleado")
    ):
        messages.error(request, "No tienes permiso para ver esta factura.")
        return redirect("index")
    return render(
        request,
        "acceso/factura_detalle.html",
        {
            "factura": factura,
            "pedido": pedido,
            "detalles": pedido.detallepedido_set.select_related("idProducto"),
        },
    )


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_compras(request):
    from .models import Pedido

    q = request.GET.get("q", "")
    pedidos = Pedido.objects.select_related("idUsuario", "sede", "factura").order_by(
        "-fechaPedido"
    )
    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(factura__numeroFactura__icontains=q)
            | Q(sede__nombre__icontains=q)
        )
    return render(
        request,
        "acceso/panel/panel_compras.html",
        {"pedidos": pedidos, "q": q, "total": pedidos.count()},
    )


@login_required
def empleado_compras(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from .models import Pedido

    q = request.GET.get("q", "")
    pedidos = (
        Pedido.objects.select_related("idUsuario", "sede", "factura")
        .prefetch_related("detallepedido_set__idProducto")
        .order_by("-fechaPedido")
    )
    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(factura__numeroFactura__icontains=q)
        )
    return render(
        request,
        "acceso/empleado/empleado_compras.html",
        {"pedidos": pedidos, "q": q, "total": pedidos.count()},
    )


# CARGA MASIVA DE CLIENTES
@login_required
@user_passes_test(es_admin, login_url="/")
def panel_clientes_importar(request):
    import csv
    import io

    import openpyxl
    from django.contrib.auth.hashers import make_password

    from .models import Rol, Usuarios

    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        nombre_archivo = archivo.name.lower()
        filas = []
        errores = []
        creados = 0

        try:
            if nombre_archivo.endswith(".xlsx"):
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                for i, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), start=2
                ):
                    filas.append((i, row))
            elif nombre_archivo.endswith(".csv") or nombre_archivo.endswith(".txt"):
                texto = archivo.read().decode("utf-8-sig")
                reader = csv.reader(io.StringIO(texto))
                next(reader, None)
                for i, row in enumerate(reader, start=2):
                    filas.append((i, row))
            else:
                messages.error(request, "Formato no soportado. Usa .xlsx, .csv o .txt")
                return redirect("panel_clientes")

            rol_cliente = Rol.objects.filter(idRol=2).first() or Rol.objects.first()

            for i, row in filas:
                try:
                    tipo_doc = str(row[0] or "").strip()
                    num_doc = str(row[1] or "").strip()
                    nombre = str(row[2] or "").strip()
                    apellidos = str(row[3] or "").strip()
                    direccion = str(row[4] or "").strip()
                    telefono = str(row[5] or "").strip()
                    correo = str(row[6] or "").strip()
                    clave = str(row[7] or "").strip() or "123456"

                    if not num_doc or not nombre:
                        errores.append(f"Fila {i}: faltan datos obligatorios.")
                        continue
                    if Usuarios.objects.filter(numDocUsuario=num_doc).exists():
                        errores.append(f"Fila {i}: documento {num_doc} ya existe.")
                        continue

                    Usuarios.objects.create(
                        tipoDocUsuario=tipo_doc,
                        numDocUsuario=num_doc,
                        nombreUsuario=nombre,
                        apellidosUsuario=apellidos or None,
                        direccionUsuario=direccion or None,
                        telefonoUsuario=(int(telefono) if telefono.isdigit() else None),
                        correoUsuario=correo or None,
                        claveUsuario=make_password(clave),
                        estadoUsuario="A",
                        idRol=rol_cliente,
                    )
                    creados += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")

            if creados:
                messages.success(
                    request,
                    f"{creados} cliente(s) importado(s) correctamente.",
                )
            for e in errores:
                messages.warning(request, e)

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect("panel_clientes")


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_categorias_pdf(request):
    from django.db.models import Count
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import CategoriaProducto

    q = request.GET.get("q", "")
    categorias = CategoriaProducto.objects.annotate(
        num_productos=Count("producto")
    ).order_by("nombreCategoria")
    if q:
        categorias = categorias.filter(nombreCategoria__icontains=q)
    html = render_to_string(
        "acceso/pdf/reporte_categorias_pdf.html",
        {
            "categorias": categorias,
            "fecha": timezone.now(),
            "total": categorias.count(),
            "q": q,
        },
    )
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_categorias.pdf"'
    return response


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_clientes_pdf(request):
    from django.db.models import Count
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import Usuarios

    q = request.GET.get("q", "")
    clientes = Usuarios.objects.annotate(num_pedidos=Count("pedido")).order_by(
        "nombreUsuario"
    )
    if q:
        clientes = clientes.filter(
            Q(nombreUsuario__icontains=q)
            | Q(correoUsuario__icontains=q)
            | Q(numDocUsuario__icontains=q)
        )
    html = render_to_string(
        "acceso/pdf/reporte_clientes_pdf.html",
        {
            "clientes": clientes,
            "fecha": timezone.now(),
            "total": clientes.count(),
            "q": q,
        },
    )
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_clientes.pdf"'
    return response


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_sedes_pdf(request):
    from django.db.models import Count
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import Sede

    q = request.GET.get("q", "")
    sedes = Sede.objects.annotate(num_pedidos=Count("pedido")).order_by("nombre")
    if q:
        sedes = sedes.filter(
            Q(nombre__icontains=q) | Q(ciudad__icontains=q) | Q(direccion__icontains=q)
        )
    html = render_to_string(
        "acceso/pdf/reporte_sedes_pdf.html",
        {
            "sedes": sedes,
            "fecha": timezone.now(),
            "total": sedes.count(),
            "q": q,
        },
    )
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_sedes.pdf"'
    return response


# CAMPAÑA MASIVA DE CORREOS
@login_required
@user_passes_test(es_admin, login_url="/")
def panel_campana(request):
    from django.contrib.auth.models import User
    from django.template.loader import render_to_string

    from .models import Usuarios

    enviados = 0
    fallidos = 0
    resultado = None

    if request.method == "POST":
        asunto = request.POST.get("asunto", "").strip()
        mensaje = request.POST.get("mensaje", "").strip()
        destinatario = request.POST.get("destinatario", "clientes")

        if not asunto or not mensaje:
            messages.error(request, "El asunto y el mensaje son obligatorios.")
        else:
            correos = []

            if destinatario == "clientes":
                correos = list(
                    Usuarios.objects.filter(
                        correoUsuario__isnull=False, estadoUsuario="A"
                    )
                    .exclude(correoUsuario="")
                    .values_list("correoUsuario", flat=True)
                )

            elif destinatario == "empleados":
                correos = list(
                    User.objects.filter(
                        empleado__isnull=False,
                        is_staff=False,
                        email__isnull=False,
                    )
                    .exclude(email="")
                    .values_list("email", flat=True)
                )

            elif destinatario == "admins":
                correos = list(
                    User.objects.filter(is_staff=True, email__isnull=False)
                    .exclude(email="")
                    .values_list("email", flat=True)
                )

            elif destinatario == "todos":
                clientes = list(
                    Usuarios.objects.filter(
                        correoUsuario__isnull=False, estadoUsuario="A"
                    )
                    .exclude(correoUsuario="")
                    .values_list("correoUsuario", flat=True)
                )
                emp_adm = list(
                    User.objects.filter(email__isnull=False)
                    .exclude(email="")
                    .values_list("email", flat=True)
                )
                correos = list(set(clientes + emp_adm))

            html_campana = render_to_string(
                "acceso/emails/campana.html",
                {
                    "asunto": asunto,
                    "mensaje": mensaje,
                },
            )

            for correo in correos:
                ok = _enviar_mailersend(correo, correo, asunto, mensaje, html_campana)
                if ok:
                    enviados += 1
                else:
                    fallidos += 1

            resultado = {
                "enviados": enviados,
                "fallidos": fallidos,
                "total": len(correos),
                "destinatario": destinatario,
            }
            if enviados > 0:
                messages.success(
                    request, f"✅ Correo enviado a {enviados} destinatario(s)."
                )
            if fallidos > 0:
                messages.warning(
                    request, f"⚠️ {fallidos} correo(s) no pudieron enviarse."
                )

    from .models import Usuarios

    total_clientes = (
        Usuarios.objects.filter(correoUsuario__isnull=False, estadoUsuario="A")
        .exclude(correoUsuario="")
        .count()
    )
    total_empleados = (
        User.objects.filter(empleado__isnull=False, is_staff=False)
        .exclude(email="")
        .count()
    )
    total_admins = User.objects.filter(is_staff=True).exclude(email="").count()

    return render(
        request,
        "acceso/panel/panel_campana.html",
        {
            "resultado": resultado,
            "total_clientes": total_clientes,
            "total_empleados": total_empleados,
            "total_admins": total_admins,
            "total_todos": total_clientes + total_empleados + total_admins,
        },
    )


# CARGA MASIVA DE PRODUCTOS (DASHBOARD)
def _safe_int(val, default=0):
    try:
        return (
            int(float(str(val)))
            if val is not None and str(val).strip() != ""
            else default
        )
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0.0):
    try:
        return (
            float(str(val)) if val is not None and str(val).strip() != "" else default
        )
    except (ValueError, TypeError):
        return default


def _descargar_imagen(url, nombre_producto, errores_list):
    if not url or not url.strip().startswith("http"):
        return None, None
    try:
        import requests
        from django.core.files.base import ContentFile

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        }
        resp = requests.get(url.strip(), timeout=10, headers=headers)
        if resp.status_code == 200:
            content_type = resp.headers.get("Content-Type", "")
            if "jpeg" in content_type or "jpg" in content_type:
                ext = "jpg"
            elif "png" in content_type:
                ext = "png"
            elif "webp" in content_type:
                ext = "webp"
            elif "gif" in content_type:
                ext = "gif"
            else:
                url_path = url.split("?")[0]
                ext = url_path.split(".")[-1][:4].lower() or "jpg"
            nombre_img = f"{nombre_producto.replace(' ', '_').lower()[:50]}.{ext}"
            return nombre_img, ContentFile(resp.content)
        else:
            errores_list.append(
                f"Imagen no descargada (HTTP {resp.status_code}): {url[:60]}"
            )
    except Exception as e:
        errores_list.append(f"Error descargando imagen: {e} — URL: {url[:60]}")
    return None, None


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_productos_importar(request):
    import csv
    import io

    import openpyxl

    from .models import CategoriaProducto, Producto

    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        nombre_archivo = archivo.name.lower()
        filas = []
        errores = []
        creados = 0
        actualizados = 0

        try:
            if nombre_archivo.endswith(".xlsx"):
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active

                header_row = None
                all_rows = list(ws.iter_rows(values_only=True))
                for idx, row in enumerate(all_rows):
                    first = str(row[0] or "").strip().lower()
                    if first == "nombreproducto":
                        header_row = idx
                        break

                if header_row is None:
                    header_row = 0

                for i, row in enumerate(
                    all_rows[header_row + 1 :], start=header_row + 2
                ):
                    if all(cell is None or str(cell).strip() == "" for cell in row):
                        continue
                    filas.append((i, row))

            elif nombre_archivo.endswith(".csv") or nombre_archivo.endswith(".txt"):
                texto = archivo.read().decode("utf-8-sig")
                reader = csv.reader(io.StringIO(texto))
                next(reader, None)
                for i, row in enumerate(reader, start=2):
                    if any(cell.strip() for cell in row):
                        filas.append((i, row))
            else:
                messages.error(request, "Formato no soportado. Usa .xlsx o .csv")
                return redirect("panel_dashboard")

            for i, row in filas:
                try:
                    nombre = str(row[0] or "").strip()
                    precio = row[1]
                    stock = row[2]
                    estado = str(row[3] or "Disponible").strip()
                    desc = str(row[4] or "").strip()
                    cat_nombre = str(row[5] or "").strip()
                    imagen_url = str(row[6] or "").strip() if len(row) > 6 else ""

                    if not nombre:
                        errores.append(f"Fila {i}: nombre obligatorio.")
                        continue

                    precio_val = _safe_float(precio, 0.0)
                    stock_val = _safe_int(stock, 0)

                    categoria = None
                    if cat_nombre:
                        categoria, _ = CategoriaProducto.objects.get_or_create(
                            nombreCategoria=cat_nombre,
                            defaults={"descripcion": ""},
                        )

                    producto_qs = Producto.objects.filter(nombreProducto=nombre)
                    if producto_qs.exists():
                        p = producto_qs.first()
                        p.precioProducto = (
                            precio_val if precio is not None else p.precioProducto
                        )
                        p.stock = stock_val if stock is not None else p.stock
                        p.estadoProducto = estado or p.estadoProducto
                        p.descripcion = desc or p.descripcion
                        if categoria:
                            p.idCategoria_Producto = categoria
                        if imagen_url and not p.imagen:
                            nombre_img, contenido = _descargar_imagen(
                                imagen_url, nombre, errores
                            )
                            if nombre_img and contenido:
                                p.imagen.save(nombre_img, contenido, save=False)
                        p.save()
                        actualizados += 1
                    else:
                        nuevo = Producto(
                            nombreProducto=nombre,
                            precioProducto=precio_val,
                            stock=stock_val,
                            estadoProducto=estado or "Disponible",
                            descripcion=desc or "",
                            idCategoria_Producto=categoria,
                        )
                        if imagen_url:
                            nombre_img, contenido = _descargar_imagen(
                                imagen_url, nombre, errores
                            )
                            if nombre_img and contenido:
                                nuevo.imagen.save(nombre_img, contenido, save=False)
                        nuevo.save()
                        creados += 1

                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")

            if creados:
                messages.success(
                    request,
                    f"✅ {creados} producto(s) creado(s) correctamente.",
                )
            if actualizados:
                messages.info(request, f"🔄 {actualizados} producto(s) actualizado(s).")
            for e in errores:
                messages.warning(request, e)

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect("panel_dashboard")


@login_required
@user_passes_test(es_admin, login_url="/")
def descargar_plantilla_productos(request):
    import csv

    from django.http import HttpResponse

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="plantilla_productos.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(
        [
            "nombreProducto",
            "precioProducto",
            "stock",
            "estadoProducto",
            "descripcion",
            "categoria",
            "imagen_url",
        ]
    )
    writer.writerow(
        [
            "Carburador VM22",
            "150000",
            "10",
            "Disponible",
            "Carburador de alto rendimiento",
            "Motor",
            "https://ejemplo.com/imagen.jpg",
        ]
    )
    writer.writerow(
        [
            "Llanta 90/90-18",
            "85000",
            "5",
            "Disponible",
            "Llanta delantera",
            "Llantas",
            "",
        ]
    )
    return response


# PERFIL DE USUARIO
@login_required
def mi_perfil(request):
    from django.contrib.auth import update_session_auth_hash

    from .models import Usuarios

    try:
        usuario_extra = Usuarios.objects.get(correoUsuario=request.user.email)
    except Usuarios.DoesNotExist:
        usuario_extra = None

    if request.method == "POST":
        accion = request.POST.get("accion")

        if accion == "cambiar_correo":
            nuevo_correo = request.POST.get("nuevo_correo", "").strip()
            password_confirm = request.POST.get("password_confirm", "")

            if not request.user.check_password(password_confirm):
                messages.error(request, "La contraseña ingresada es incorrecta.")
            elif not nuevo_correo:
                messages.error(request, "El correo no puede estar vacío.")
            elif (
                request.user.__class__.objects.filter(email=nuevo_correo)
                .exclude(pk=request.user.pk)
                .exists()
            ):
                messages.error(request, "Ese correo ya está en uso por otra cuenta.")
            else:
                request.user.email = nuevo_correo
                request.user.save()
                if usuario_extra:
                    usuario_extra.correoUsuario = nuevo_correo
                    usuario_extra.save()
                messages.success(request, "¡Correo actualizado correctamente!")

        elif accion == "cambiar_password":
            password_actual = request.POST.get("password_actual", "")
            password_nuevo = request.POST.get("password_nuevo", "")
            password_nuevo2 = request.POST.get("password_nuevo2", "")

            if not request.user.check_password(password_actual):
                messages.error(request, "La contraseña actual es incorrecta.")
            elif len(password_nuevo) < 6:
                messages.error(
                    request,
                    "La nueva contraseña debe tener al menos 6 caracteres.",
                )
            elif password_nuevo != password_nuevo2:
                messages.error(request, "Las contraseñas nuevas no coinciden.")
            else:
                request.user.set_password(password_nuevo)
                request.user.save()
                if usuario_extra:
                    from django.contrib.auth.hashers import make_password

                    usuario_extra.claveUsuario = make_password(password_nuevo)
                    usuario_extra.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, "¡Contraseña cambiada correctamente!")

    return render(
        request,
        "acceso/perfil.html",
        {
            "usuario_extra": usuario_extra,
        },
    )


# ── PANEL VENTAS ──────────────────────────────────────────────────────────────

@login_required
@user_passes_test(es_admin, login_url="/")
def panel_ventas(request):
    from django.db.models import Avg, Count, Sum
    from django.utils import timezone

    from .models import DetallePedido, Pedido

    q             = request.GET.get("q", "")
    estado_filtro = request.GET.get("estado", "")
    metodo_filtro = request.GET.get("metodo", "")
    fecha_desde   = request.GET.get("fecha_desde", "")
    fecha_hasta   = request.GET.get("fecha_hasta", "")

    pedidos = (
        Pedido.objects
        .select_related("idUsuario", "sede", "factura")
        .prefetch_related("detallepedido_set__idProducto")
        .order_by("-fechaPedido")
    )

    total_ventas      = pedidos.count()
    total_ingresos    = pedidos.aggregate(t=Sum("totalPedido"))["t"] or 0
    ticket_promedio   = pedidos.aggregate(a=Avg("totalPedido"))["a"] or 0
    ventas_entregadas = pedidos.filter(estadoPedido="Entregado").count()
    hoy               = timezone.now().date()
    ventas_hoy        = pedidos.filter(fechaPedido__date=hoy).count()
    ingresos_hoy      = pedidos.filter(fechaPedido__date=hoy).aggregate(t=Sum("totalPedido"))["t"] or 0

    top_productos = (
        DetallePedido.objects
        .values("idProducto__nombreProducto")
        .annotate(total_vendido=Sum("cantidad"))
        .order_by("-total_vendido")[:5]
    )

    metodos_pago  = (
        Pedido.objects.values("metodoPago")
        .annotate(cantidad=Count("idPedido"), total=Sum("totalPedido"))
        .order_by("-total")
    )
    todos_metodos = Pedido.objects.values_list("metodoPago", flat=True).distinct()

    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(idUsuario__apellidosUsuario__icontains=q)
            | Q(factura__numeroFactura__icontains=q)
            | Q(ciudad__icontains=q)
        )
    if estado_filtro:
        pedidos = pedidos.filter(estadoPedido=estado_filtro)
    if metodo_filtro:
        pedidos = pedidos.filter(metodoPago=metodo_filtro)
    if fecha_desde:
        pedidos = pedidos.filter(fechaPedido__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos = pedidos.filter(fechaPedido__date__lte=fecha_hasta)

    total_filtrado          = pedidos.count()
    total_ingresos_filtrado = pedidos.aggregate(t=Sum("totalPedido"))["t"] or 0

    return render(request, "acceso/panel/panel_ventas.html", {
        "total_ventas":             total_ventas,
        "total_ingresos":           total_ingresos,
        "ticket_promedio":          ticket_promedio,
        "ventas_entregadas":        ventas_entregadas,
        "ventas_hoy":               ventas_hoy,
        "ingresos_hoy":             ingresos_hoy,
        "top_productos":            top_productos,
        "metodos_pago":             metodos_pago,
        "todos_metodos":            todos_metodos,
        "pedidos":                  pedidos,
        "total_filtrado":           total_filtrado,
        "total_ingresos_filtrado":  total_ingresos_filtrado,
        "q":                        q,
        "estado_filtro":            estado_filtro,
        "metodo_filtro":            metodo_filtro,
        "fecha_desde":              fecha_desde,
        "fecha_hasta":              fecha_hasta,
        # ✅ AGREGADO: lista de estados para el select del filtro
        "estados_choices": [
            "Procesado", "Confirmado", "En Proceso",
            "Listo", "Entregado", "Cancelado",
        ],
    })


@login_required
@user_passes_test(es_admin, login_url="/")
def panel_venta_detalle(request, pk):
    from .models import Pedido

    pedido = get_object_or_404(
        Pedido.objects
        .select_related("idUsuario", "sede", "factura")
        .prefetch_related("detallepedido_set__idProducto__idCategoria_Producto"),
        pk=pk,
    )
    if request.method == "POST":
        pedido.estadoPedido = request.POST.get("estadoPedido", pedido.estadoPedido)
        pedido.save()
        messages.success(request, f"Estado de la venta #{pk} actualizado.")
        return redirect("panel_venta_detalle", pk=pk)

    estados = ["Procesado", "Confirmado", "En Proceso", "Listo", "Entregado", "Cancelado"]

    return render(request, "acceso/panel/panel_venta_detalle.html", {
        "pedido": pedido,
        "estados": estados,
    })


@login_required
@user_passes_test(es_admin, login_url="/")
def reporte_ventas_pdf(request):
    from django.db.models import Avg, Sum
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone

    from .models import Pedido

    q             = request.GET.get("q", "")
    estado_filtro = request.GET.get("estado", "")
    metodo_filtro = request.GET.get("metodo", "")
    fecha_desde   = request.GET.get("fecha_desde", "")
    fecha_hasta   = request.GET.get("fecha_hasta", "")

    pedidos = Pedido.objects.select_related("idUsuario", "factura").order_by("-fechaPedido")

    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(idUsuario__apellidosUsuario__icontains=q)
            | Q(factura__numeroFactura__icontains=q)
            | Q(ciudad__icontains=q)
        )
    if estado_filtro:
        pedidos = pedidos.filter(estadoPedido=estado_filtro)
    if metodo_filtro:
        pedidos = pedidos.filter(metodoPago=metodo_filtro)
    if fecha_desde:
        pedidos = pedidos.filter(fechaPedido__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos = pedidos.filter(fechaPedido__date__lte=fecha_hasta)

    partes = []
    if q:             partes.append(f"Búsqueda: {q}")
    if estado_filtro: partes.append(f"Estado: {estado_filtro}")
    if metodo_filtro: partes.append(f"Método: {metodo_filtro}")
    if fecha_desde:   partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta:   partes.append(f"Hasta: {fecha_hasta}")

    html_string = render_to_string("acceso/pdf/reporte_ventas_pdf.html", {
        "pedidos":           pedidos,
        "total_ventas":      pedidos.count(),
        "total_ingresos":    pedidos.aggregate(t=Sum("totalPedido"))["t"] or 0,
        "ticket_promedio":   pedidos.aggregate(a=Avg("totalPedido"))["a"] or 0,
        "ventas_entregadas": pedidos.filter(estadoPedido="Entregado").count(),
        "fecha_reporte":     timezone.now().strftime("%d/%m/%Y %H:%M"),
        "filtros":           " | ".join(partes),
    })

    buffer = _generar_pdf(html_string)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_ventas.pdf"'
    return response

@login_required
def empleado_ventas(request):
    if not hasattr(request.user, 'empleado'):
        return redirect('login_empleado')
    from django.db.models import Avg, Sum
    from .models import Pedido
 
    q             = request.GET.get('q', '')
    estado_filtro = request.GET.get('estado', '')
    fecha_desde   = request.GET.get('fecha_desde', '')
    fecha_hasta   = request.GET.get('fecha_hasta', '')
 
    pedidos = Pedido.objects.select_related('idUsuario', 'factura').order_by('-fechaPedido')
 
    total_ventas      = pedidos.count()
    total_ingresos    = pedidos.aggregate(t=Sum('totalPedido'))['t'] or 0
    ticket_promedio   = pedidos.aggregate(a=Avg('totalPedido'))['a'] or 0
    ventas_entregadas = pedidos.filter(estadoPedido='Entregado').count()
 
    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(idUsuario__apellidosUsuario__icontains=q)
            | Q(factura__numeroFactura__icontains=q)
            | Q(ciudad__icontains=q)
        )
    if estado_filtro:
        pedidos = pedidos.filter(estadoPedido=estado_filtro)
    if fecha_desde:
        pedidos = pedidos.filter(fechaPedido__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos = pedidos.filter(fechaPedido__date__lte=fecha_hasta)
 
    return render(request, 'acceso/empleado/empleado_ventas.html', {
        'pedidos':                pedidos,
        'total_ventas':           total_ventas,
        'total_ingresos':         total_ingresos,
        'ticket_promedio':        ticket_promedio,
        'ventas_entregadas':      ventas_entregadas,
        'total_filtrado':         pedidos.count(),
        'total_ingresos_filtrado': pedidos.aggregate(t=Sum('totalPedido'))['t'] or 0,
        'q':                      q,
        'estado_filtro':          estado_filtro,
        'fecha_desde':            fecha_desde,
        'fecha_hasta':            fecha_hasta,
        'estados_choices': ['Procesado', 'Confirmado', 'En Proceso', 'Listo', 'Entregado', 'Cancelado'],
    })
 
 
@login_required
def empleado_facturas(request):
    if not hasattr(request.user, 'empleado'):
        return redirect('login_empleado')
    from .models import Factura
 
    q           = request.GET.get('q', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
 
    facturas = (
        Factura.objects
        .select_related('idPedido__idUsuario', 'idPedido__sede')
        .order_by('-fechaFactura')
    )
 
    if q:
        facturas = facturas.filter(
            Q(numeroFactura__icontains=q)
            | Q(idPedido__idUsuario__nombreUsuario__icontains=q)
            | Q(idPedido__idUsuario__apellidosUsuario__icontains=q)
        )
    if fecha_desde:
        facturas = facturas.filter(fechaFactura__date__gte=fecha_desde)
    if fecha_hasta:
        facturas = facturas.filter(fechaFactura__date__lte=fecha_hasta)
 
    return render(request, 'acceso/empleado/empleado_facturas.html', {
        'facturas':    facturas,
        'total':       facturas.count(),
        'q':           q,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })

@login_required
def empleado_venta_detalle(request, pk):
    if not hasattr(request.user, 'empleado'):
        return redirect('login_empleado')
    from .models import Pedido
 
    pedido = get_object_or_404(
        Pedido.objects
        .select_related("idUsuario", "sede", "factura")
        .prefetch_related("detallepedido_set__idProducto__idCategoria_Producto"),
        pk=pk,
    )
    
    if request.method == "POST":
        pedido.estadoPedido = request.POST.get("estadoPedido", pedido.estadoPedido)
        pedido.save()
        messages.success(request, f"Estado de la venta #{pk} actualizado.")
        return redirect("empleado_venta_detalle", pk=pk)
 
    estados = ["Procesado", "Confirmado", "En Proceso", "Listo", "Entregado", "Cancelado"]
 
    return render(request, "acceso/empleado/empleado_venta_detalle.html", {
        "pedido": pedido,
        "estados": estados,
    })

@login_required
def empleado_reporte_productos_pdf(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from django.db.models import Count
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone
    from .models import CategoriaProducto, Producto
 
    q = request.GET.get("q", "")
    categoria_id = request.GET.get("categoria", "")
    estado = request.GET.get("estado", "")
 
    productos = Producto.objects.select_related("idCategoria_Producto").order_by("nombreProducto")
    if q:
        productos = productos.filter(nombreProducto__icontains=q)
    if categoria_id:
        productos = productos.filter(idCategoria_Producto__idCategoria_Producto=categoria_id)
    if estado == "Agotado":
        productos = productos.filter(stock=0)
    elif estado == "stock_bajo":
        productos = productos.filter(stock__gt=0, stock__lte=5)
    elif estado:
        productos = productos.filter(estadoProducto=estado)
 
    html = render_to_string("acceso/empleado/pdf/empleado_reporte_productos_pdf.html", {
        "productos": productos,
        "fecha": timezone.now(),
        "total": productos.count(),
        "q": q,
    })
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_productos.pdf"'
    return response
 
 
@login_required
def empleado_reporte_categorias_pdf(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from django.db.models import Count
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone
    from .models import CategoriaProducto
 
    q = request.GET.get("q", "")
    categorias = CategoriaProducto.objects.annotate(num_productos=Count("producto")).order_by("nombreCategoria")
    if q:
        categorias = categorias.filter(
            Q(nombreCategoria__icontains=q) | Q(descripcion__icontains=q)
        )
 
    html = render_to_string("acceso/empleado/pdf/empleado_reporte_categorias_pdf.html", {
        "categorias": categorias,
        "fecha": timezone.now(),
        "total": categorias.count(),
        "q": q,
    })
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_categorias.pdf"'
    return response
 
 
@login_required
def empleado_reporte_pedidos_pdf(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone
    from .models import Pedido
 
    q = request.GET.get("q", "")
    estado = request.GET.get("estado", "")
    sede_id = request.GET.get("sede", "")
 
    pedidos = Pedido.objects.select_related("idUsuario", "sede").order_by("-fechaPedido")
    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(idUsuario__apellidosUsuario__icontains=q)
            | Q(ciudad__icontains=q)
        )
    if estado:
        pedidos = pedidos.filter(estadoPedido=estado)
    if sede_id:
        pedidos = pedidos.filter(sede_id=sede_id)
 
    partes = []
    if q:       partes.append(f"Búsqueda: {q}")
    if estado:  partes.append(f"Estado: {estado}")
 
    html = render_to_string("acceso/empleado/pdf/empleado_reporte_pedidos_pdf.html", {
        "pedidos": pedidos,
        "fecha": timezone.now(),
        "total": pedidos.count(),
        "filtros": " | ".join(partes),
    })
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_pedidos.pdf"'
    return response
 
 
@login_required
def empleado_reporte_ventas_pdf(request):
    if not hasattr(request.user, "empleado"):
        return redirect("login_empleado")
    from django.db.models import Avg, Sum
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.utils import timezone
    from .models import Pedido
 
    q = request.GET.get("q", "")
    estado_filtro = request.GET.get("estado", "")
    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")
 
    pedidos = Pedido.objects.select_related("idUsuario", "factura").order_by("-fechaPedido")
    if q:
        pedidos = pedidos.filter(
            Q(idUsuario__nombreUsuario__icontains=q)
            | Q(idUsuario__apellidosUsuario__icontains=q)
            | Q(factura__numeroFactura__icontains=q)
            | Q(ciudad__icontains=q)
        )
    if estado_filtro:
        pedidos = pedidos.filter(estadoPedido=estado_filtro)
    if fecha_desde:
        pedidos = pedidos.filter(fechaPedido__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos = pedidos.filter(fechaPedido__date__lte=fecha_hasta)
 
    partes = []
    if q:             partes.append(f"Búsqueda: {q}")
    if estado_filtro: partes.append(f"Estado: {estado_filtro}")
    if fecha_desde:   partes.append(f"Desde: {fecha_desde}")
    if fecha_hasta:   partes.append(f"Hasta: {fecha_hasta}")
 
    html = render_to_string("acceso/empleado/pdf/empleado_reporte_ventas_pdf.html", {
        "pedidos": pedidos,
        "total_ventas": pedidos.count(),
        "total_ingresos": pedidos.aggregate(t=Sum("totalPedido"))["t"] or 0,
        "ticket_promedio": pedidos.aggregate(a=Avg("totalPedido"))["a"] or 0,
        "ventas_entregadas": pedidos.filter(estadoPedido="Entregado").count(),
        "fecha_reporte": timezone.now().strftime("%d/%m/%Y %H:%M"),
        "filtros": " | ".join(partes),
    })
    buffer = _generar_pdf(html)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_ventas_empleado.pdf"'
    return response

@login_required
@user_passes_test(es_admin, login_url="/")
def panel_cliente_toggle(request, pk):
    from .models import Usuarios

    if request.method == "POST":
        cliente = get_object_or_404(Usuarios, pk=pk)
        if cliente.estadoUsuario == "A":
            cliente.estadoUsuario = "I"
            messages.success(request, f"{cliente.nombreUsuario} ha sido inactivado.")
        else:
            cliente.estadoUsuario = "A"
            messages.success(request, f"{cliente.nombreUsuario} ha sido activado.")
        cliente.save()
    return redirect("panel_clientes")
